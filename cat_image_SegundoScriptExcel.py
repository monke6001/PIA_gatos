# 🔧 Instalación automática de librerías si no están presentes
import importlib
import subprocess
import sys

def instalar_si_no(paq):
    if importlib.util.find_spec(paq) is None:
        print(f"📦 Instalando '{paq}'...")
        subprocess.check_call([sys.executable, "-m", "pip", "install", paq])
    else:
        print(f"✅ Librería '{paq}' ya está instalada.")

# Lista de paquetes requeridos
paquetes_requeridos = ["pandas", "openpyxl", "matplotlib"]

for paquete in paquetes_requeridos:
    instalar_si_no(paquete)

# Ahora importa el resto


import json
import os
import re
import pandas as pd
import matplotlib.pyplot as plt
from collections import Counter
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

def load_json_data(filename):
    if not os.path.exists(filename):
        print(f"❌ El archivo {filename} no existe.")
        return []
    try:
        with open(filename, "r", encoding="utf-8") as f:
            data = json.load(f)
        print(f"✅ Archivo '{filename}' cargado con éxito. Total registros: {len(data)}")
        return data
    except json.JSONDecodeError:
        print("❌ Error al decodificar el archivo JSON.")
        return []

def validate_data(data):
    pattern_url = re.compile(r"^https?://")
    pattern_text = re.compile(r"^[\w\s\-,'áéíóúüñÁÉÍÓÚÜÑ]+$")

    valid_data = []
    for item in data:
        if (
            pattern_url.match(item.get("url", "")) and
            pattern_text.match(item.get("breed", "")) and
            pattern_text.match(item.get("origin", "")) and
            pattern_text.match(item.get("temperament", ""))
        ):
            valid_data.append(item)
    return valid_data

def show_full_data(data):
    if not data:
        print("⚠️ No hay datos para mostrar.")
        return
    print("\n📋 Mostrando todos los gatos disponibles:\n")
    for idx, cat in enumerate(data, start=1):
        print(f"Gato #{idx}")
        print(f"• URL         : {cat.get('url', 'N/A')}")
        print(f"• Raza        : {cat.get('breed', 'N/A')}")
        print(f"• Origen      : {cat.get('origin', 'N/A')}")
        print(f"• Temperamento: {cat.get('temperament', 'N/A')}")
        print("-" * 50)

def show_summary(data):
    breeds = [item["breed"] for item in data]
    origins = [item["origin"] for item in data]
    print("\n📊 Razas más comunes:")
    for breed, count in Counter(breeds).most_common():
        print(f"• {breed}: {count}")
    print("\n🌍 Orígenes más comunes:")
    for origin, count in Counter(origins).most_common():
        print(f"• {origin}: {count}")

def filter_by_origin(data, origin_name):
    filtered = [item for item in data if item["origin"].lower() == origin_name.lower()]
    if not filtered:
        print(f"⚠️ No se encontraron gatos del origen '{origin_name}'.")
    else:
        print(f"\n🐾 Gatos del origen '{origin_name}':")
        for idx, cat in enumerate(filtered, start=1):
            print(f"\nGato #{idx}")
            print(f"• URL         : {cat['url']}")
            print(f"• Raza        : {cat['breed']}")
            print(f"• Origen      : {cat['origin']}")
            print(f"• Temperamento: {cat['temperament']}")
            print("-" * 50)

def format_sheet(sheet):
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows(min_row=1, max_row=1):
        for cell in row:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = thin_border
    for col in sheet.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value)
                if val:
                    max_length = max(max_length, len(val))
            except:
                pass
        sheet.column_dimensions[col_letter].width = min(max_length + 2, 40)

def create_graphs(data):
    df = pd.DataFrame(data)
    breed_counts = df['breed'].value_counts()
    origin_counts = df['origin'].value_counts()

    plt.figure(figsize=(10, 5))
    plt.plot(breed_counts.values[:10], marker='o')
    plt.title("Top 10 Razas Más Comunes")
    plt.xlabel("Índice")
    plt.ylabel("Cantidad")
    plt.grid(True)
    plt.savefig("grafico_lineas.png")
    plt.close()

    plt.figure(figsize=(10, 5))
    breed_counts[:10].plot(kind="bar", color="skyblue")
    plt.title("Top 10 Razas (Barras)")
    plt.xlabel("Raza")
    plt.ylabel("Cantidad")
    plt.xticks(rotation=45)
    plt.grid(True)
    plt.tight_layout()
    plt.savefig("grafico_barras.png")
    plt.close()

    plt.figure(figsize=(10, 5))
    plt.scatter(range(len(origin_counts[:10])), origin_counts.values[:10], c='green')
    plt.title("Distribución de Orígenes")
    plt.xlabel("Índice")
    plt.ylabel("Cantidad")
    plt.grid(True)
    plt.savefig("grafico_dispersion.png")
    plt.close()

    plt.figure(figsize=(8, 8))
    origin_counts[:5].plot(kind="pie", autopct='%1.1f%%', startangle=90)
    plt.title("Distribución de Orígenes (Top 5)")
    plt.ylabel("")
    plt.savefig("grafico_pastel.png")
    plt.close()

def insert_graphs_into_excel(excel_file):
    wb = load_workbook(excel_file)
    graph_sheet = wb.create_sheet("Gráficas")
    positions = [("grafico_lineas.png", "B2"),
                 ("grafico_barras.png", "L2"),
                 ("grafico_dispersion.png", "B30"),
                 ("grafico_pastel.png", "L30")]
    for file, position in positions:
        if os.path.exists(file):
            img = ExcelImage(file)
            img.width = 500
            img.height = 300
            graph_sheet.add_image(img, position)
    wb.save(excel_file)
    print("🖼️ Gráficas embellecidas e insertadas correctamente.")

def open_excel_file(file_path):
    if os.name == 'nt':
        os.system(f'start EXCEL.EXE "{file_path}"')
    elif os.name == 'posix':
        os.system(f'open "{file_path}"')

def export_to_excel(data, filename="gatos_exportados.xlsx"):
    if not data:
        print("⚠️ No hay datos para exportar.")
        return
    df = pd.DataFrame(data)
    breed_counts = df['breed'].value_counts().reset_index()
    breed_counts.columns = ['Raza', 'Cantidad']
    origin_counts = df['origin'].value_counts().reset_index()
    origin_counts.columns = ['Origen', 'Cantidad']

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="Datos Completos", index=False)
        breed_counts.to_excel(writer, sheet_name="Resumen Razas", index=False)
        origin_counts.to_excel(writer, sheet_name="Resumen Orígenes", index=False)

    wb = load_workbook(filename)
    for sheet_name in wb.sheetnames:
        if sheet_name != "Gráficas":
            format_sheet(wb[sheet_name])
    wb.save(filename)

    create_graphs(data)
    insert_graphs_into_excel(filename)
    open_excel_file(filename)

# ---------- EJECUCIÓN PRINCIPAL ----------
if __name__ == "__main__":
    print("📂 Analizador de Datos de Gatos (desde JSON)")
    data = load_json_data("cat_data_clean.json")

    if data:
        data = validate_data(data)
        show_full_data(data)
        show_summary(data)

        origin_input = input("\n🔍 Filtrar por país de origen (opcional): ").strip()
        if origin_input:
            filter_by_origin(data, origin_input)

        export_to_excel(data)
